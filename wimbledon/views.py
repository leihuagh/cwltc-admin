import calendar
from django.shortcuts import redirect
from django.views.generic import TemplateView
from django.http import HttpResponse
from braces.views import LoginRequiredMixin, StaffuserRequiredMixin
from openpyxl import Workbook
from .models import *

class BallotView(LoginRequiredMixin, TemplateView):
    template_name = 'wimbledon/ballot.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['bg_class'] = 'bg-white'
        return context

class ChoiceView(LoginRequiredMixin, TemplateView):
    template_name = 'wimbledon/choice.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return common_context(context, self.request)

    def post(self, request, *args, **kwargs):
        if 'cancel' in request.POST:
            return redirect('wimbledon_ballot')
        person = Person.objects.get(auth=request.user.id)
        OptOut.objects.filter(person_id=person.id).delete()
        if 'all' in request.POST:
            opt = OptOut.objects.create(person_id=person.id, ticket_id=None, all_days=True)
        else:
            for key in request.POST:
                if key[:6] == 'ticket':
                    id = key[7:]
                    opt = OptOut.objects.create(person_id=person.id, ticket_id=id)
        return redirect('wimbledon_done')


class DoneView(LoginRequiredMixin, TemplateView):
    template_name = 'wimbledon/done.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return common_context(context, self.request)

def common_context(context, request):
    person = Person.objects.get(auth=request.user.id)
    all_days = OptOut.objects.filter(all_days=True)
    if len(all_days) > 0:
        context['all_days'] = True
        context['opt_out'] = []
    else:
        context['all_days'] = False
        context['opt_out'] = list(OptOut.objects.filter(person_id=person.id).values_list('ticket_id', flat=True))
    context['tickets'] = list(Tickets.objects.order_by('id'))
    context['bg_class'] = 'bg-white'
    return context


class ExportView(StaffuserRequiredMixin, TemplateView):
    template_name= 'wimbledon/export.html'

    def post(self, request):

        """ https://djangotricks.blogspot.co.uk/2013/12/how-to-export-data-as-excel.html """
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Wimbledon opt out.xlsx'
        wb = Workbook()
        ws = wb.active
        ws.title = 'Wimbledon Opt Out'

        data = OptOut.objects.all().order_by('person__first_name', 'ticket__day')
        row = 1
        for dat in data:
            person = Person.objects.get(pk=dat.person_id)
            ws.cell(row, 1, person.fullname)
            if dat.all_days:
                ws.cell(row, 2, 'Opt out')
            else:
                ticket = Tickets.objects.get(pk=dat.ticket_id)
                ws.cell(row, 3, ticket.court)
                ws.cell(row, 4, ticket.day)
                ws.cell(row, 5, calendar.day_name[ticket.date.weekday()])
                ws.cell(row, 6, ticket.date)

            row += 1
        wb.save(response)
        return response