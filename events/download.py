from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from events.models import Event, Participant


def export_tournament(tournament):
    """ Export all events in a tournament"""
    wb = Workbook()
    events = Event.objects.filter(tournament=tournament, active=True).order_by('pk')
    count = 0
    for event in events:
        if count == 0:
            ws = wb.get_active_sheet()
            ws.title = event.name
        else:
            ws = wb.create_sheet(title=event.name)
        count += 1
        dump_event(event, ws)
    return save_response(wb, tournament.name)


def export_event(event):
    """ Export a single event """
    wb = Workbook()
    ws = wb.active
    ws.title = event.name
    dump_event(event, ws)
    return save_response(wb, event.name)


def save_response(workbook, name):
    """ save workbook and create Http response """
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={name}.xlsx'
    workbook.save(response)
    return response


def dump_event(event, worksheet):
    """ Dump single event to a worksheet """
    participants = Participant.objects.filter(event=event)
    columns = [('Id', 10), ('Person', 40), ]
    if event.with_partner():
        columns += [('Id2', 10), ('Partner', 40), ('Combined', 60)]

    for col_num in range(len(columns)):
        c = worksheet.cell(1, col_num + 1)
        c.value = columns[col_num][0]
        c.font = Font(sz=12, bold=True)
        worksheet.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]

    row_num = 1
    for part in participants:
        row_num += 1
        row = [part.person_id, part.person.fullname]
        if event.with_partner():
            if part.partner:
                row += [part.partner_id, part.partner.fullname, f'{part.person.fullname} & {part.partner.fullname}']
            else:
                row +=[0, 'No partner']
        worksheet.append(row)
