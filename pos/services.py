import datetime
import logging
from decimal import Decimal
from django.db import transaction
from django.db.models import Sum
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from .models import Transaction, LineItem, Layout, PosPayment, Item, Location, TWO_PLACES
from members.models import InvoiceItem, ItemType
from pos.models import VisitorBook

stdlogger = logging.getLogger(__name__)


class Error(Exception):
    """
    Base class for exceptions in this module
    """
    pass


class PosServicesError(Error):
    """
    Error while processing payment
    """
    def __init__(self, message):
        self.message = message


def unbilled_transactions_total(person, item_type):
    return Transaction.objects.filter(
        person=person,
        billed=False,
        item_type=item_type
    ).aggregate(Sum('total'))


@transaction.atomic
def create_transaction_from_receipt(creator_id, terminal, layout_id, receipt, total,
                                    people, attended, creation_date=None):
    """
    Create Transaction, LineItem and PosPayment records in the database
    Return a description of it
    """
    try:
        complimentary = False
        count = len(people)
        dec_total = (Decimal(total)/100).quantize(TWO_PLACES)
        item_type = Layout.objects.get(pk=layout_id).item_type
        if count > 0:
            person_id = int(people[0]['id'])
            if person_id == -1:
                complimentary = True
                person_id = None
        else:
            person_id = None
        trans = Transaction(
            creation_date=creation_date,
            creator_id=creator_id,
            person_id=person_id,
            terminal=terminal,
            item_type=item_type,
            total=dec_total,
            billed=False,
            cash=person_id == None and not complimentary,
            complimentary=complimentary,
            split=count > 1,
            attended=attended
            )
        trans.save()

        for item_dict in receipt:
            line_item = LineItem(
                item_id=item_dict['id'],
                sale_price=Decimal(item_dict['sale_price']).quantize(TWO_PLACES),
                cost_price=Decimal(item_dict['cost_price']).quantize(TWO_PLACES),
                quantity=1,
                transaction=trans
                )
            line_item.save()

        if complimentary:
            return ('Complimentary', dec_total)
        if trans.cash:
            return ('Cash', dec_total)

        pay_total = Decimal(0)
        for person in people:
            pos_payment = PosPayment(
                transaction=trans,
                person_id=person['id'],
                billed=False,
                total=(Decimal(person['amount'])/100).quantize(TWO_PLACES)
            )
            pay_total += pos_payment.total
            pos_payment.save()
        if pay_total !=  dec_total:
            stdlogger.error(f'ERROR: POS Transaction total: {dec_total} unequal to Payment total: {pay_total} Id: {trans.id}')
        return (people[0]['name'], dec_total)
    except Exception as e:
        raise PosServicesError('Error creating transaction')


def delete_billed_transactions(before_date):
    """
    Delete transactions that have been billed and linked items and payments
    """
    trans = Transaction.objects.filter(billed=True, creation_date__lt=before_date)
    count = trans.count()
    trans.delete()
    return count


def delete_billed_visitors(before_date):
    """
    Delete visitor book entries that have been billed
    """
    visitors = VisitorBook.objects.filter(billed=True, creation_date__lt=before_date)
    count = visitors.count()
    visitors.delete()
    return count


def dump_items_to_excel(item_type_id):
    """ https://djangotricks.blogspot.co.uk/2013/12/how-to-export-data-as-excel.html """
    queryset= Item.objects.filter(item_type_id=item_type_id)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Items.xlsx'

    wb = Workbook()
    ws = wb.active
    ws.title = 'Items'

    row_num = 0

    columns = [
        ('Description', 40),
        ('Price', 10),
    ]

    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        c.font = Font(sz=12, bold=True)
        ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]

    for obj in queryset:
        row_num += 1
        row = [
            obj.description,
            obj.sale_price,
        ]
        for col_num in range(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
            if col_num == 1:
                c.number_format = '£0.00'

    wb.save(response)
    return response


def dump_layout_to_excel(layout):
    """ https://djangotricks.blogspot.co.uk/2013/12/how-to-export-data-as-excel.html """
    array, items = build_pos_array(layout)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Price list.xlsx'

    wb = Workbook()
    ws = wb.active
    ws.title = 'Price List'

    widths = [10, 40, 10]

    for col_num in range(len(widths)):
        ws.column_dimensions[get_column_letter(col_num + 1)].width = widths[col_num]

    c = ws.cell(row=1, column=2, value='Price List')

    row_num = 2
    for row in array:
        for col_num in range(len(row)):
            if col_num == 0:
                if len(row[col_num]) > 2:
                    description = row[col_num][2]
                    ws.cell(row=row_num, column=1, value=description)
                    row_num += 1
            else:
                if len(row[col_num]) > 2:
                    item = row[col_num][2]
                    ws.cell(row=row_num, column=2, value=item.description)
                    c = ws.cell(row=row_num, column=3, value=item.sale_price)
                    c.number_format = '£0.00'
                    row_num += 1
    wb.save(response)
    return response


def build_pos_array(layout=None):
    """
    Build an array of rows and columns
    Col[0] is the description for a row
    Cells will contain items
    Returns the used items for managing the layout
    """
    rows = []
    for r in range(1, Location.ROW_MAX + 1):
        cols = []
        for c in range(0, Location.COL_MAX + 1):
            cols.append([r, c])
        rows.append(cols)
    items = None
    if layout: # true when managing a layout
        locations = Location.objects.filter(layout_id=layout.id).order_by('row', 'col').prefetch_related('item').prefetch_related('item__colour')
        items = Item.objects.filter(item_type_id=layout.item_type_id).order_by('button_text')
        for loc in locations:
            if loc.col == 0:
                rows[loc.row - 1][loc.col].append(loc.description)
            else:
                rows[loc.row - 1][loc.col].append(loc.item)
                if items:
                    item = [item for item in items if item.button_text == loc.item.button_text]
                    if item:
                        item[0].used = True
    return rows, items